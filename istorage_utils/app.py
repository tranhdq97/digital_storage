# using flask_restful
from flask import Flask, jsonify, request, send_file, after_this_request
from flask_restful import Resource, Api
from flask_cors import CORS

# pyopenxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side

# utilities
import os
import random
import string
import argparse


MINE_TYPE_FOR_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


FONT_DOCUMENT_HEADLINE = Font(name='Times New Roman', size=14, bold=True)
FONT_DOCUMENT_HEADLINE_UNDERLINED = Font(
    name='Times New Roman',
    size=13, bold=True,
    underline='single'
)
FONT_DOCUMENT_HEADLINE_NOT_BOLD = Font(
    name='Times New Roman',
    size=14,
    bold=False
)
FONT_HEADER_ITALIC = Font(name='Times New Roman',
                          size=13, italic=True, bold=True)
FONT_HEADER_ITALIC_NOT_BOLD = Font(name='Times New Roman',
                                   size=13, italic=True, bold=False)
FONT_TABLE_HEADLINE = Font(name='Times New Roman', size=13, bold=True)
FONT_TABLE_CELL = Font(name='Times New Roman', size=13, bold=False)


ALIGNMENT_CENTER = Alignment(horizontal='center', vertical='center')
ALIGNMENT_CENTER_WRAP_TEXT = Alignment(wrap_text=True)


TABLE_BORDER_SIDE = Side(border_style='thick', color='000000')
TABLE_BORDER = Border(
    left=TABLE_BORDER_SIDE,
    right=TABLE_BORDER_SIDE,
    top=TABLE_BORDER_SIDE,
    bottom=TABLE_BORDER_SIDE
)


def get_random_string(length):
    # choose from all lowercase letter
    letters = string.ascii_lowercase
    result_str = ''.join(random.choice(letters) for i in range(length))
    return result_str


def get_data_and_delete_file(file_name):
    file = open(file_name, "rb")
    data = file.read()
    file.close()
    os.remove(file_name)  # clean tmp file
    return data


def make_mucluc(xl_file_name, data):
    wb = Workbook()
    ws = wb.active

    ws.merge_cells('A1:F1')
    ws.merge_cells('A2:F2')

    ws['A1'].font = FONT_DOCUMENT_HEADLINE_NOT_BOLD
    ws['A1'].alignment = ALIGNMENT_CENTER
    ws['A1'] = 'UBND TỈNH QUẢNG NGÃI'

    ws['A2'].font = FONT_DOCUMENT_HEADLINE_UNDERLINED
    ws['A2'].alignment = ALIGNMENT_CENTER
    ws['A2'] = 'SỞ THÔNG TIN VÀ TRUYỀN THÔNG'

    ws.merge_cells('J1:P1')
    ws.merge_cells('J2:P2')

    ws['J1'].font = FONT_DOCUMENT_HEADLINE
    ws['J1'].alignment = ALIGNMENT_CENTER
    ws['J1'] = 'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM'

    ws['J2'].font = FONT_DOCUMENT_HEADLINE_UNDERLINED
    ws['J2'].alignment = ALIGNMENT_CENTER
    ws['J2'] = 'Độc lập - Tự do - Hạnh phúc'

    ws.merge_cells('A3:P4')
    ws.merge_cells('A5:P5')
    ws['A5'].font = FONT_DOCUMENT_HEADLINE
    ws['A5'].alignment = ALIGNMENT_CENTER
    ws['A5'] = 'MỤC LỤC HỒ SƠ, TÀI LIỆU NỘP LƯU'

    ws.merge_cells('A6:P6')
    ws['A6'].font = FONT_DOCUMENT_HEADLINE
    ws['A6'].alignment = ALIGNMENT_CENTER
    ws['A6'] = 'Thời hạn bảo quản: Vĩnh viễn'

    ws.merge_cells('A7:P7')
    ws['A7'].font = FONT_DOCUMENT_HEADLINE_UNDERLINED
    ws['A7'].alignment = ALIGNMENT_CENTER
    ws['A7'] = 'Năm: 2023'

    ws.merge_cells('A8:P9')

    start_row_of_table = 10
    len_data = len(data)
    end_row_of_table = start_row_of_table + len_data

    # Set border for table before merging any pair of cells
    for idx in range(start_row_of_table, end_row_of_table + 1):
        for char_idx in range(ord('A'), ord('P') + 1):
            char_val = chr(char_idx)
            ws[char_val + str(idx)].border = TABLE_BORDER

    ws.merge_cells('B' + str(start_row_of_table) +
                   ':D' + str(start_row_of_table))
    ws.merge_cells('E' + str(start_row_of_table) +
                   ':I' + str(start_row_of_table))
    ws.merge_cells('J' + str(start_row_of_table) +
                   ':K' + str(start_row_of_table))
    ws.merge_cells('L' + str(start_row_of_table) +
                   ':M' + str(start_row_of_table))
    ws.merge_cells('O' + str(start_row_of_table) +
                   ':P' + str(start_row_of_table))

    ws['A' + str(start_row_of_table)] = "Số TT"
    ws['B' + str(start_row_of_table)] = "Số, ký hiệu hồ sơ"
    ws['E' + str(start_row_of_table)] = "Tiêu đề hồ sơ"
    ws['J' + str(start_row_of_table)] = "Thời gian tài liệu"
    ws['L' + str(start_row_of_table)] = "Thời hạn bảo quản"
    ws['N' + str(start_row_of_table)] = "Số tờ / số trang"
    ws['O' + str(start_row_of_table)] = "Ghi chú"

    ws['A' + str(start_row_of_table)].font = FONT_TABLE_HEADLINE
    ws['B' + str(start_row_of_table)].font = FONT_TABLE_HEADLINE
    ws['E' + str(start_row_of_table)].font = FONT_TABLE_HEADLINE
    ws['J' + str(start_row_of_table)].font = FONT_TABLE_HEADLINE
    ws['L' + str(start_row_of_table)].font = FONT_TABLE_HEADLINE
    ws['N' + str(start_row_of_table)].font = FONT_TABLE_HEADLINE
    ws['O' + str(start_row_of_table)].font = FONT_TABLE_HEADLINE

    ws['A' + str(start_row_of_table)].alignment = ALIGNMENT_CENTER
    ws['B' + str(start_row_of_table)].alignment = ALIGNMENT_CENTER
    ws['E' + str(start_row_of_table)].alignment = ALIGNMENT_CENTER
    ws['J' + str(start_row_of_table)].alignment = ALIGNMENT_CENTER
    ws['L' + str(start_row_of_table)].alignment = ALIGNMENT_CENTER
    ws['N' + str(start_row_of_table)].alignment = ALIGNMENT_CENTER
    ws['O' + str(start_row_of_table)].alignment = ALIGNMENT_CENTER

    for idx in range(len_data):
        row_id = idx + start_row_of_table + 1
        current_file = data[idx]

        ws.merge_cells('B' + str(row_id) +
                       ':D' + str(row_id))
        ws.merge_cells('E' + str(row_id) +
                       ':I' + str(row_id))
        ws.merge_cells('J' + str(row_id) +
                       ':K' + str(row_id))
        ws.merge_cells('L' + str(row_id) +
                       ':M' + str(row_id))
        ws.merge_cells('O' + str(row_id) +
                       ':P' + str(row_id))

        ws['A' + str(row_id)] = idx + 1
        ws['B' + str(row_id)] = current_file.get("gov_file_code", "")
        ws['E' + str(row_id)] = current_file.get("title", "")
        ws['J' + str(row_id)] = current_file.get("start_date", "")
        ws['L' + str(row_id)] = current_file.get("maintenance", "")
        ws['N' + str(row_id)] = current_file.get("page_number", "")
        ws['O' + str(row_id)] = current_file.get("note", "")

        ws['A' + str(row_id)].font = FONT_TABLE_CELL
        ws['B' + str(row_id)].font = FONT_TABLE_CELL
        ws['E' + str(row_id)].font = FONT_TABLE_CELL
        ws['J' + str(row_id)].font = FONT_TABLE_CELL
        ws['L' + str(row_id)].font = FONT_TABLE_CELL
        ws['N' + str(row_id)].font = FONT_TABLE_CELL
        ws['O' + str(row_id)].font = FONT_TABLE_CELL

    ws.merge_cells('A' + str(end_row_of_table + 1) +
                   ':P' + str(end_row_of_table + 2))

    start_row_of_footer = end_row_of_table + 3

    ws.merge_cells('A' + str(start_row_of_footer) +
                   ':P' + str(start_row_of_footer))
    ws['A' + str(start_row_of_footer)
       ] = "Mục lục này gồm: ............... hồ sơ (đơn vị bảo quản)"
    ws['A' + str(start_row_of_footer)].font = FONT_HEADER_ITALIC
    ws['A' + str(start_row_of_footer)].alignment = ALIGNMENT_CENTER

    ws.merge_cells('A' + str(start_row_of_footer + 1) +
                   ':P' + str(start_row_of_footer + 1))
    ws['A' + str(start_row_of_footer + 1)
       ] = "Viết bằng chữ: ............... hồ sơ (đơn vị bảo quản)"
    ws['A' + str(start_row_of_footer + 1)].font = FONT_HEADER_ITALIC
    ws['A' + str(start_row_of_footer + 1)].alignment = ALIGNMENT_CENTER

    ws.merge_cells('A' + str(start_row_of_footer + 2) +
                   ':P' + str(start_row_of_footer + 3))

    ws.merge_cells('L' + str(start_row_of_footer + 4) +
                   ':P' + str(start_row_of_footer + 4))
    ws['L' + str(start_row_of_footer + 4)
       ] = ".............., ngày ..... tháng ..... năm ......."
    ws['L' + str(start_row_of_footer + 4)].font = FONT_HEADER_ITALIC
    ws['L' + str(start_row_of_footer + 4)].alignment = ALIGNMENT_CENTER

    ws.merge_cells('A' + str(start_row_of_footer + 5) +
                   ':P' + str(start_row_of_footer + 5))

    ws.merge_cells('L' + str(start_row_of_footer + 6) +
                   ':P' + str(start_row_of_footer + 6))
    ws['L' + str(start_row_of_footer + 6)] = "Người lập"
    ws['L' + str(start_row_of_footer + 6)].font = FONT_HEADER_ITALIC_NOT_BOLD
    ws['L' + str(start_row_of_footer + 6)].alignment = ALIGNMENT_CENTER

    ws.merge_cells('L' + str(start_row_of_footer + 7) +
                   ':P' + str(start_row_of_footer + 7))
    ws['L' + str(start_row_of_footer + 7)] = "(Ký và ghi rõ họ tên, chức vụ)"
    ws['L' + str(start_row_of_footer + 7)].font = FONT_HEADER_ITALIC
    ws['L' + str(start_row_of_footer + 7)].alignment = ALIGNMENT_CENTER

    wb.save(xl_file_name)


def make_danhmuc(xl_file_name):
    pass


def make_mucluc_vanban(xl_file_name):
    pass


# creating the flask app
app = Flask(__name__)
# Enable CORS
CORS(app)
# creating an API object
api = Api(app)

# making a class for a particular resource
# the get, post methods correspond to get and post requests
# they are automatically mapped by flask_restful.
# other methods include put, delete, etc.


class Hello(Resource):
    # corresponds to the GET request.
    # this function is called whenever there
    # is a GET request for this resource
    def get(self):
        return jsonify({'message': 'hello world!!'})

    # Corresponds to POST request
    def post(self):
        data = request.get_json()	 # status code
        return jsonify({'data': data}), 201


# another resource to calculate the square of a number
class Square(Resource):
    def get(self, num):
        return jsonify({'square': num**2})


class Excel(Resource):
    def post(self):
        post_data = request.get_json()

        def generate():
            xl_file_name = get_random_string(10) + ".xlsx"
            make_mucluc(xl_file_name, post_data["data"])

            data = get_data_and_delete_file(xl_file_name)
            yield data

        return app.response_class(generate(), mimetype=MINE_TYPE_FOR_XLSX)


# adding the defined resources along with their corresponding urls
api.add_resource(Hello, '/')
api.add_resource(Square, '/square/<int:num>')
api.add_resource(Excel, '/excel')


def make_parser():
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--ip", default="0.0.0.0",
                        help="Ip of running host")
    parser.add_argument("-p", "--port", default=4444, help="Port of host")
    return parser


args = make_parser().parse_args()

# driver function
if __name__ == '__main__':
    app.run(debug=True, host=args.ip, port=args.port)
